# -*- coding: utf-8 -*-
"""
controllers/pase_lista_pivot.py
================================
Endpoints JSON para la vista matricial interactiva del pase de lista.

GET  /actividades/pase-lista/<int:actividad_id>/data
     → Devuelve la estructura pivot (fechas, filas, readonly).

POST /actividades/pase-lista/toggle
     → body: { asistencia_id: int, presente: bool }
     → Actualiza el campo `presente` del registro indicado.

GET  /actividades/pase-lista/<int:actividad_id>/exportar-excel
     → Descarga un archivo .xlsx con el pase de lista completo.
"""

import io
import json
import logging
import re
import traceback
import unicodedata
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from odoo import http
from odoo.http import request

_logger = logging.getLogger(__name__)


class PaseListaPivotController(http.Controller):

    # ------------------------------------------------------------------
    # GET /actividades/pase-lista/<actividad_id>/data
    # ------------------------------------------------------------------

    @http.route(
        "/actividades/pase-lista/<int:actividad_id>/data",
        type="http",
        auth="user",
        methods=["GET"],
        csrf=False,
    )
    def get_pivot_data(self, actividad_id, **kw):
        """Devuelve JSON con la estructura matricial del pase de lista."""
        pivot_model = request.env["actividad.pase.lista.pivot"]
        data = pivot_model.get_pivot_data(actividad_id)
        return request.make_response(
            json.dumps(data),
            headers=[("Content-Type", "application/json")],
        )

    # ------------------------------------------------------------------
    # POST /actividades/pase-lista/toggle
    # ------------------------------------------------------------------

    @http.route(
        "/actividades/pase-lista/toggle",
        type="json",
        auth="user",
        methods=["POST"],
        csrf=False,
    )
    def toggle_asistencia(self, asistencia_id, presente, **kw):
        """Actualiza presente/ausente de un registro de asistencia."""
        rec = request.env["actividad.asistencia"].browse(int(asistencia_id))
        if not rec.exists():
            return {"ok": False, "error": "Registro no encontrado."}

        actividad = rec.actividad_id
        if actividad.estado_code == "finalizada" or actividad.certificates_generated:
            return {"ok": False, "error": "La actividad no permite modificaciones."}

        rec.write({"presente": bool(presente)})
        return {"ok": True, "presente": rec.presente}

    # ------------------------------------------------------------------
    # GET /actividades/pase-lista/<actividad_id>/exportar-excel
    # ------------------------------------------------------------------

    @http.route(
        "/actividades/pase-lista/<int:actividad_id>/exportar-excel",
        type="http",
        auth="user",
        methods=["GET"],
        csrf=False,
    )
    def exportar_excel(self, actividad_id, **kw):
        """Genera y descarga un .xlsx con el pase de lista completo."""
        try:
            return self._generar_excel(actividad_id)
        except Exception as exc:
            _logger.error("exportar_excel error: %s", traceback.format_exc())
            return request.make_response(
                f"Error al generar el Excel: {exc}\n\n{traceback.format_exc()}",
                status=500,
                headers=[("Content-Type", "text/plain; charset=utf-8")],
            )

    def _generar_excel(self, actividad_id):
        env = request.env
        Asistencia = env["actividad.asistencia"]
        actividad = env["actividad.complementaria"].browse(actividad_id)
        if not actividad.exists():
            return request.make_response(
                "Actividad no encontrada.",
                headers=[("Content-Type", "text/plain")],
                status=404,
            )

        # ── Recopilar datos ──────────────────────────────────────────
        records = Asistencia.search(
            [("actividad_id", "=", actividad_id)],
            order="fecha asc, partner_id asc",
        )

        fechas = sorted({r.fecha for r in records})
        fechas_str = [str(f) for f in fechas]

        # Inscripciones ordenadas alfabéticamente
        inscripciones = actividad.inscripcion_ids.sorted(
            key=lambda i: i.partner_id.name or ""
        )

        # Mapa (inscripcion_id, fecha) → presente
        mapa = {}
        for r in records:
            mapa[(r.inscripcion_id.id, str(r.fecha))] = r.presente

        # ── Estilos ──────────────────────────────────────────────────
        AZUL_OSCURO = "1F3864"
        AZUL_CLARO = "D6E4F7"
        VERDE = "C6EFCE"
        ROJO = "FFCCCC"
        GRIS = "F2F2F2"
        BORDE = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def estilo_encabezado(celda, color_fondo=AZUL_OSCURO, color_fuente="FFFFFF"):
            celda.font = Font(name="Arial", bold=True, color=color_fuente, size=10)
            celda.fill = PatternFill("solid", start_color=color_fondo)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = BORDE

        def estilo_dato(celda, fondo=None):
            celda.font = Font(name="Arial", size=10)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = BORDE
            if fondo:
                celda.fill = PatternFill("solid", start_color=fondo)

        # ── Workbook ─────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pase de Lista"

        # Fila 1: título de la actividad (fusionada)
        total_cols = 2 + len(fechas) + 1   # Estudiante + fechas + Total
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        titulo = ws.cell(row=1, column=1, value=f"Pase de Lista — {actividad.name}")
        titulo.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        titulo.fill = PatternFill("solid", start_color=AZUL_OSCURO)
        titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Fila 2: periodo / fecha de exportación
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        subtitulo_txt = f"Exportado el {date.today().strftime('%d/%m/%Y')}"
        if actividad.fecha_inicio and actividad.fecha_fin:
            subtitulo_txt = (
                f"Periodo: {actividad.fecha_inicio.strftime('%d/%m/%Y')} – "
                f"{actividad.fecha_fin.strftime('%d/%m/%Y')}   |   {subtitulo_txt}"
            )
        subtitulo = ws.cell(row=2, column=1, value=subtitulo_txt)
        subtitulo.font = Font(name="Arial", italic=True, size=9, color="444444")
        subtitulo.fill = PatternFill("solid", start_color=AZUL_CLARO)
        subtitulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16

        # Fila 3: encabezados de columna
        HEADER_ROW = 3
        ws.cell(row=HEADER_ROW, column=1, value="#")
        estilo_encabezado(ws.cell(row=HEADER_ROW, column=1))

        ws.cell(row=HEADER_ROW, column=2, value="Estudiante")
        estilo_encabezado(ws.cell(row=HEADER_ROW, column=2))

        meses = ["ene", "feb", "mar", "abr", "may", "jun",
                 "jul", "ago", "sep", "oct", "nov", "dic"]
        for col_idx, fiso in enumerate(fechas_str, start=3):
            partes = fiso.split("-")
            label = f"{int(partes[2])} {meses[int(partes[1])-1]}"
            c = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
            estilo_encabezado(c)

        total_col = 2 + len(fechas) + 1
        c_total = ws.cell(row=HEADER_ROW, column=total_col, value="Asistencias")
        estilo_encabezado(c_total)
        ws.row_dimensions[HEADER_ROW].height = 32

        # Filas de datos
        for fila_idx, insc in enumerate(inscripciones, start=1):
            row = HEADER_ROW + fila_idx
            fondo_fila = None if fila_idx % 2 == 1 else GRIS

            # Número
            c_num = ws.cell(row=row, column=1, value=fila_idx)
            estilo_dato(c_num, fondo_fila)

            # Nombre (alineado a la izquierda)
            c_nom = ws.cell(row=row, column=2, value=insc.partner_id.name or "(Sin nombre)")
            c_nom.font = Font(name="Arial", size=10)
            c_nom.alignment = Alignment(horizontal="left", vertical="center")
            c_nom.border = BORDE
            if fondo_fila:
                c_nom.fill = PatternFill("solid", start_color=fondo_fila)

            # Celdas de asistencia
            total_presente = 0
            for col_idx, fiso in enumerate(fechas_str, start=3):
                presente = mapa.get((insc.id, fiso))
                if presente is True:
                    valor = "✓"
                    fondo = VERDE
                    total_presente += 1
                elif presente is False:
                    valor = "✗"
                    fondo = ROJO
                else:
                    valor = "—"
                    fondo = fondo_fila

                c = ws.cell(row=row, column=col_idx, value=valor)
                estilo_dato(c, fondo)
                c.font = Font(
                    name="Arial",
                    bold=(presente is not None),
                    size=11,
                    color="1A7A1A" if presente else ("C00000" if presente is False else "888888"),
                )

            # Total
            c_tot = ws.cell(row=row, column=total_col,
                            value=f"{total_presente}/{len(fechas)}")
            estilo_dato(c_tot, fondo_fila)
            c_tot.font = Font(name="Arial", bold=True, size=10)

        # Fila de totales por sesión
        DATA_ROWS = len(inscripciones)
        totals_row = HEADER_ROW + DATA_ROWS + 1
        ws.merge_cells(start_row=totals_row, start_column=1,
                       end_row=totals_row, end_column=2)
        c_lbl = ws.cell(row=totals_row, column=1, value="Total presentes")
        estilo_encabezado(c_lbl, color_fondo="2E4057")
        for col_idx, fiso in enumerate(fechas_str, start=3):
            total = sum(
                1 for insc in inscripciones
                if mapa.get((insc.id, fiso)) is True
            )
            c = ws.cell(row=totals_row, column=col_idx, value=total)
            estilo_encabezado(c, color_fondo="2E4057")
        ws.cell(row=totals_row, column=total_col, value="")
        estilo_encabezado(ws.cell(row=totals_row, column=total_col), color_fondo="2E4057")

        # ── Anchos de columna ────────────────────────────────────────
        ws.column_dimensions["A"].width = 5    # #
        ws.column_dimensions["B"].width = 32   # Estudiante
        for col_idx in range(3, 3 + len(fechas)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
        ws.column_dimensions[get_column_letter(total_col)].width = 12

        # Fijar encabezados al hacer scroll
        ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=3)

        # ── Serializar y devolver ────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Normalizar: quitar acentos, espacios → guion bajo, minúsculas
        nombre_raw = actividad.name or "actividad"
        nombre_norm = unicodedata.normalize("NFD", nombre_raw)
        nombre_ascii = "".join(
            c for c in nombre_norm
            if unicodedata.category(c) != "Mn"
        )
        nombre_slug = re.sub(r"[^a-zA-Z0-9]+", "_", nombre_ascii).strip("_").lower()
        nombre_archivo = f"{nombre_slug}_lista.xlsx"
        return request.make_response(
            buf.read(),
            headers=[
                ("Content-Type",
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="{nombre_archivo}"'),
                ("Access-Control-Expose-Headers", "Content-Disposition"),
            ],
        )
