# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import ValidationError


class VentanaLiberacion(models.Model):
    """SE-02SC: Período habilitado para recibir y evaluar solicitudes de liberación."""

    _name = 'actividad.ventana.liberacion'
    _description = 'Ventana de Liberación de Créditos Complementarios'
    _order = 'fecha_inicio desc'
    _rec_name = 'display_name'

    periodo_id = fields.Many2one(
        'sii.periodo', string='Periodo Escolar', required=True, ondelete='restrict',
    )
    fecha_inicio = fields.Date(string='Fecha de Inicio', required=True)
    fecha_fin = fields.Date(string='Fecha de Cierre', required=True)
    activa = fields.Boolean(
        string='Activa', compute='_compute_activa', store=True,
    )
    reporte_generado = fields.Boolean(
        string='Reporte Generado', default=False, readonly=True, copy=False,
    )
    reporte_attachment_id = fields.Many2one(
        'ir.attachment', string='Reporte Excel', readonly=True, copy=False,
    )
    display_name = fields.Char(compute='_compute_display_name', store=True)

    @api.depends('periodo_id', 'fecha_inicio', 'fecha_fin')
    def _compute_display_name(self):
        for rec in self:
            if rec.periodo_id and rec.fecha_inicio and rec.fecha_fin:
                rec.display_name = (
                    f'{rec.periodo_id.clave_periodo} '
                    f'({rec.fecha_inicio.strftime("%d/%m/%Y")} – '
                    f'{rec.fecha_fin.strftime("%d/%m/%Y")})'
                )
            else:
                rec.display_name = 'Nueva ventana'

    @api.depends('fecha_inicio', 'fecha_fin')
    def _compute_activa(self):
        from datetime import date
        hoy = date.today()
        for rec in self:
            rec.activa = bool(
                rec.fecha_inicio and rec.fecha_fin
                and rec.fecha_inicio <= hoy <= rec.fecha_fin
            )

    # ── Constraints ──────────────────────────────────────────────────────────

    @api.constrains('fecha_inicio', 'fecha_fin', 'periodo_id')
    def _check_fechas(self):
        for rec in self:
            if rec.fecha_fin <= rec.fecha_inicio:
                raise ValidationError(
                    'La fecha de cierre debe ser posterior a la fecha de inicio.'
                )
            if rec.periodo_id:
                p = rec.periodo_id
                if rec.fecha_inicio < p.fecha_inicio or rec.fecha_fin > p.fecha_fin:
                    raise ValidationError(
                        'Las fechas de la ventana deben estar dentro del periodo escolar '
                        f'({p.fecha_inicio} – {p.fecha_fin}).'
                    )
            # RN2: no overlap con otra ventana activa del mismo periodo
            solapada = self.search([
                ('periodo_id', '=', rec.periodo_id.id),
                ('id', '!=', rec._origin.id or 0),
                ('fecha_inicio', '<=', rec.fecha_fin),
                ('fecha_fin', '>=', rec.fecha_inicio),
            ], limit=1)
            if solapada:
                raise ValidationError(
                    f'Ya existe una ventana que se superpone en el mismo periodo: '
                    f'{solapada.display_name}.'
                )

    # ── Business logic ────────────────────────────────────────────────────────

    def action_generar_reporte(self):
        """SE-02SC RN5: genera el reporte Excel de estudiantes liberados."""
        self.ensure_one()
        pendientes = self.env['actividad.solicitud.liberacion'].search_count([
            ('ventana_id', '=', self.id),
            ('estado', '=', 'en_revision'),
        ])
        if pendientes:
            raise ValidationError(
                f'Aún quedan {pendientes} solicitud(es) sin evaluar en esta ventana. '
                'Evalúalas antes de generar el reporte.'
            )
        aprobadas = self.env['actividad.solicitud.liberacion'].search([
            ('ventana_id', '=', self.id),
            ('estado', '=', 'aprobada'),
        ])
        if not aprobadas:
            raise ValidationError('No hay solicitudes aprobadas en esta ventana.')

        import io
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ValidationError('openpyxl no está instalado en el entorno.')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Estudiantes Liberados'

        # Encabezados
        headers = [
            'Nombre del Aprobador', 'Nombre del Estudiante',
            'No. Control', 'Nivel de Desempeño', 'Semestre',
        ]
        header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row_idx, sol in enumerate(aprobadas, 2):
            estudiante = sol.estudiante_id
            # Buscar datos en sii.estudiante por login
            sii_est = self.env['sii.estudiante'].sudo().search(
                [('no_control', '=', estudiante.login)], limit=1
            )
            semestre = sii_est.semestre if sii_est else ''
            no_control = sii_est.no_control if sii_est else estudiante.login
            ws.append([
                sol.aprobado_por.name if sol.aprobado_por else '',
                estudiante.name,
                no_control,
                sol.promedio_label,
                semestre,
            ])

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import base64
        nombre = (
            f'ReporteLiberados_{self.periodo_id.clave_periodo}_'
            f'{self.fecha_fin.strftime("%Y%m%d")}.xlsx'
        )
        attachment = self.env['ir.attachment'].sudo().create({
            'name': nombre,
            'type': 'binary',
            'datas': base64.b64encode(buf.read()).decode(),
            'mimetype': (
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
            'res_model': self._name,
            'res_id': self.id,
        })
        self.write({'reporte_generado': True, 'reporte_attachment_id': attachment.id})
        access_token = attachment.sudo().generate_access_token()[0]
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?access_token={access_token}&download=true',
            'target': 'self',
        }
